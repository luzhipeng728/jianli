"""简历导出服务 - 支持JSON/XML/Excel格式"""
import io
import json
from typing import List
from xml.etree.ElementTree import Element, SubElement, tostring
from xml.dom import minidom
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from app.models.resume import ResumeData


class ResumeExporter:
    """简历导出器"""

    def to_json(self, resumes: List[ResumeData], indent: int = 2) -> str:
        """导出为JSON格式"""
        data = [resume.model_dump(mode="json") for resume in resumes]
        return json.dumps(data, ensure_ascii=False, indent=indent)

    def to_xml(self, resumes: List[ResumeData]) -> str:
        """导出为XML格式"""
        root = Element("resumes")
        root.set("count", str(len(resumes)))

        for resume in resumes:
            resume_elem = SubElement(root, "resume")
            resume_elem.set("id", resume.id)

            # 基本信息
            basic = SubElement(resume_elem, "basic_info")
            self._add_text_element(basic, "name", resume.basic_info.name)
            self._add_text_element(basic, "phone", resume.basic_info.phone)
            self._add_text_element(basic, "email", resume.basic_info.email)
            self._add_text_element(basic, "gender", resume.basic_info.gender)
            if resume.basic_info.age:
                self._add_text_element(basic, "age", str(resume.basic_info.age))

            # 教育经历
            education_elem = SubElement(resume_elem, "education")
            for edu in resume.education:
                edu_elem = SubElement(education_elem, "item")
                self._add_text_element(edu_elem, "school", edu.school)
                self._add_text_element(edu_elem, "degree", edu.degree)
                self._add_text_element(edu_elem, "major", edu.major)
                self._add_text_element(edu_elem, "start_date", edu.start_date or "")
                self._add_text_element(edu_elem, "end_date", edu.end_date or "")

            # 工作经历
            experience_elem = SubElement(resume_elem, "experience")
            for exp in resume.experience:
                exp_elem = SubElement(experience_elem, "item")
                self._add_text_element(exp_elem, "company", exp.company)
                self._add_text_element(exp_elem, "title", exp.title)
                self._add_text_element(exp_elem, "start_date", exp.start_date or "")
                self._add_text_element(exp_elem, "end_date", exp.end_date or "")
                self._add_text_element(exp_elem, "duties", exp.duties)

            # 技能
            skills_elem = SubElement(resume_elem, "skills")
            hard_skills = SubElement(skills_elem, "hard_skills")
            for skill in resume.skills.hard_skills:
                self._add_text_element(hard_skills, "skill", skill)
            soft_skills = SubElement(skills_elem, "soft_skills")
            for skill in resume.skills.soft_skills:
                self._add_text_element(soft_skills, "skill", skill)

            # 求职意向
            intention = SubElement(resume_elem, "job_intention")
            self._add_text_element(intention, "position", resume.job_intention.position)
            self._add_text_element(intention, "location", resume.job_intention.location)
            if resume.job_intention.salary_min:
                self._add_text_element(intention, "salary_min", str(resume.job_intention.salary_min))
            if resume.job_intention.salary_max:
                self._add_text_element(intention, "salary_max", str(resume.job_intention.salary_max))

            # 警告
            if resume.warnings:
                warnings_elem = SubElement(resume_elem, "warnings")
                for warning in resume.warnings:
                    warn_elem = SubElement(warnings_elem, "warning")
                    warn_elem.set("type", warning.type)
                    warn_elem.text = warning.message

            # 文件信息
            self._add_text_element(resume_elem, "file_name", resume.file_name)
            self._add_text_element(resume_elem, "file_type", resume.file_type)
            self._add_text_element(resume_elem, "created_at", resume.created_at.isoformat())

        # 格式化输出
        rough_string = tostring(root, encoding='unicode')
        reparsed = minidom.parseString(rough_string)
        return reparsed.toprettyxml(indent="  ")

    def _add_text_element(self, parent: Element, tag: str, text: str):
        """添加文本元素"""
        elem = SubElement(parent, tag)
        elem.text = text or ""

    def to_excel(self, resumes: List[ResumeData]) -> bytes:
        """导出为Excel格式"""
        wb = Workbook()

        # 样式定义
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # === 汇总表 ===
        ws_summary = wb.active
        ws_summary.title = "简历汇总"

        summary_headers = [
            "序号", "姓名", "电话", "邮箱", "性别", "年龄",
            "最高学历", "学校", "专业",
            "最近公司", "最近职位", "工作年限",
            "硬技能", "期望职位", "期望薪资", "期望地点", "警告数"
        ]

        # 写入表头
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入数据
        for row, resume in enumerate(resumes, 2):
            # 计算工作年限
            exp_years = self._calculate_experience_years(resume.experience)

            # 最高学历
            edu = resume.education[0] if resume.education else None
            exp = resume.experience[0] if resume.experience else None

            # 薪资范围
            salary = ""
            if resume.job_intention.salary_min or resume.job_intention.salary_max:
                min_s = f"{resume.job_intention.salary_min // 1000}K" if resume.job_intention.salary_min else ""
                max_s = f"{resume.job_intention.salary_max // 1000}K" if resume.job_intention.salary_max else ""
                salary = f"{min_s}-{max_s}" if min_s and max_s else (min_s or max_s)

            data = [
                row - 1,
                resume.basic_info.name,
                resume.basic_info.phone,
                resume.basic_info.email,
                resume.basic_info.gender,
                resume.basic_info.age,
                edu.degree if edu else "",
                edu.school if edu else "",
                edu.major if edu else "",
                exp.company if exp else "",
                exp.title if exp else "",
                f"{exp_years}年" if exp_years else "",
                ", ".join(resume.skills.hard_skills[:5]),
                resume.job_intention.position,
                salary,
                resume.job_intention.location,
                len(resume.warnings)
            ]

            for col, value in enumerate(data, 1):
                cell = ws_summary.cell(row=row, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border

        # 调整列宽
        column_widths = [6, 10, 15, 25, 8, 8, 10, 20, 15, 20, 15, 10, 30, 15, 15, 12, 8]
        for i, width in enumerate(column_widths, 1):
            ws_summary.column_dimensions[chr(64 + i)].width = width

        # 冻结首行
        ws_summary.freeze_panes = "A2"

        # === 详细信息表 ===
        ws_detail = wb.create_sheet("详细信息")

        detail_headers = [
            "姓名", "联系方式",
            "教育经历", "工作经历",
            "硬技能", "软技能",
            "求职意向", "风险警告"
        ]

        for col, header in enumerate(detail_headers, 1):
            cell = ws_detail.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row, resume in enumerate(resumes, 2):
            # 联系方式
            contact = f"电话: {resume.basic_info.phone}\n邮箱: {resume.basic_info.email}"

            # 教育经历
            edu_text = "\n\n".join([
                f"{e.school} | {e.degree} | {e.major}\n{e.start_date or ''} - {e.end_date or ''}"
                for e in resume.education
            ])

            # 工作经历
            exp_text = "\n\n".join([
                f"{e.company} | {e.title}\n{e.start_date or ''} - {e.end_date or ''}\n{e.duties[:200]}"
                for e in resume.experience
            ])

            # 求职意向
            intention = f"职位: {resume.job_intention.position}\n地点: {resume.job_intention.location}"
            if resume.job_intention.salary_min or resume.job_intention.salary_max:
                intention += f"\n薪资: {resume.job_intention.salary_min or 0}-{resume.job_intention.salary_max or 0}"

            # 警告
            warnings = "\n".join([f"[{w.type}] {w.message}" for w in resume.warnings])

            data = [
                resume.basic_info.name,
                contact,
                edu_text,
                exp_text,
                ", ".join(resume.skills.hard_skills),
                ", ".join(resume.skills.soft_skills),
                intention,
                warnings
            ]

            for col, value in enumerate(data, 1):
                cell = ws_detail.cell(row=row, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border

        # 调整列宽
        detail_widths = [12, 25, 40, 50, 40, 30, 25, 30]
        for i, width in enumerate(detail_widths, 1):
            ws_detail.column_dimensions[chr(64 + i)].width = width

        ws_detail.freeze_panes = "A2"

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _calculate_experience_years(self, experiences) -> int:
        """计算总工作年限"""
        total = 0
        for exp in experiences:
            try:
                start = exp.start_date or ""
                end = exp.end_date or "至今"
                start_year = int(start.split(".")[0]) if start and start[0].isdigit() else 0
                end_year = 2025 if "至今" in end else (int(end.split(".")[0]) if end and end[0].isdigit() else 0)
                if start_year and end_year:
                    total += max(0, end_year - start_year)
            except:
                pass
        return total


# 全局导出器实例
resume_exporter = ResumeExporter()
